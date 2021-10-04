from openpyxl import *
from tkinter import *

file = load_workbook(r'C:\\Users\\US\\Desktop\\Python Programming\\files\\GUI-Tkinter-data.xlsx')
sheet = file.active


def excel():
    sheet.column_dimensions['A'].width = 30
    sheet.cell(row = 1, column = 1).value = "Name"

def clear():
    name_field.delete(0,END)

def insert():
    if (name_field.get() == "" ):
        print("Please Give Input")
    else:
        current_row = sheet.max_row
        current_column = sheet.max_column

        sheet.cell(row = current_row + 1, column = 1).value = name_field.get()
        file.save(r'GUI-Tkinter-data.xlsx')
        clear()


if __name__ == "__main__":
    # Create a GUI window
    root = Tk()
    root.title("Registration Form")
    
    # Create a Name Lable
    name = Label(root, text = 'Name')
    name.grid( row = 1, column = 0)                 # Name lable's placement

    # Create a text box for typing the name
    name_field = Entry(root)
    name_field.grid(row = 1, column = 1, ipadx = '100')         # ipadx --> how much space

    # Call excel() function
    excel()

    # Create a submit button 
    submit = Button( root, text = 'Submit', command = insert)   # command --> call insert() function
    submit.grid( row = 8, column = 1)

    # Start the GUI
    root.mainloop()



